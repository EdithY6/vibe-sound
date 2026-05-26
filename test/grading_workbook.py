#!/usr/bin/env python3
"""
VibeSound batch test — grading Excel + manifest from numbered images.

Pictures: one folder, filenames = number (1.jpg, 02.png, 003.jpeg, …).

Creates / updates:
  - batch_manifest.csv          (for batch_eval.py)
  - vibesound_batch_grading.xlsx (human grading + pipeline columns)

Protocol (from rubric): 1 rater (R1), listen first 8 seconds.

Usage:
  python grading_workbook.py create --images-dir ./images --count 50
  python grading_workbook.py create --images-dir ./images --texts texts.csv
  python grading_workbook.py merge --workbook vibesound_batch_grading.xlsx \\
      --results batch_results/results.csv
"""
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

# Weights for total_score_0to100 (must match Excel formula in sheet)
WEIGHTS = {
    "mood_fit_1to5": 0.25,
    "scene_fit_1to5": 0.15,
    "coherence_1to5": 0.15,
    "reel_usability_1to5": 0.15,
    "audio_quality_1to5": 0.10,
    "distinctiveness_1to5": 0.10,
    "prompt_faithfulness_1to5": 0.10,
}

SCORE_COLS = list(WEIGHTS.keys())

HEADERS = [
    "picture_no",
    "image_file",
    "user_text",
    "scene_caption",
    "target_mood",
    "mood_score",
    "music_prompt",
    "output_audio",
    "rater_id",
    "listen_seconds",
    *SCORE_COLS,
    "would_post_YN",
    "notes",
    "total_score_0to100",
]

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp"}

# Excel column letters for formula (A=1 …)
def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s


def _stem_number(path: Path) -> str | None:
    """Extract numeric id from stem: '1', '02', '003' → sortable int string."""
    m = re.fullmatch(r"(\d+)", path.stem)
    return m.group(1) if m else None


def discover_images(images_dir: Path, *, count: int = 0) -> list[Path]:
    if not images_dir.is_dir():
        raise FileNotFoundError(f"Images folder not found: {images_dir}")
    files = [
        p
        for p in images_dir.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS and _stem_number(p)
    ]
    files.sort(key=lambda p: int(p.stem))
    if count > 0:
        files = files[:count]
    return files


def load_texts_csv(path: Path) -> dict[str, str]:
    """CSV: picture_no,user_text  OR  picture_no,text"""
    out: dict[str, str] = {}
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            key = (
                row.get("picture_no")
                or row.get("id")
                or row.get("sample_id")
                or ""
            ).strip()
            text = (row.get("user_text") or row.get("text") or "").strip()
            if key:
                out[key.lstrip("0") or "0"] = text
                out[key.zfill(len(key))] = text  # keep both forms
    return out


def rows_from_images(
    images: list[Path],
    texts: dict[str, str] | None,
    *,
    rater_id: str = "R1",
    listen_seconds: int = 8,
    audio_subdir: str = "audio",
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for p in images:
        pic_no = str(int(p.stem))  # normalise 007 → 7 for display; keep file name as-is
        stem_key = p.stem
        user_text = ""
        if texts:
            user_text = texts.get(stem_key, "") or texts.get(pic_no, "")
        rel_image = p.name
        row = {h: "" for h in HEADERS}
        row.update(
            {
                "picture_no": pic_no,
                "image_file": rel_image,
                "user_text": user_text,
                "rater_id": rater_id,
                "listen_seconds": str(listen_seconds),
                "output_audio": f"{audio_subdir}/{p.stem}.mp4",
            }
        )
        rows.append(row)
    return rows


def write_manifest(
    rows: list[dict[str, str]], images_dir: Path, out_csv: Path
) -> None:
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    with out_csv.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f, fieldnames=["sample_id", "image_path", "user_text"]
        )
        w.writeheader()
        for r in rows:
            w.writerow(
                {
                    "sample_id": r["picture_no"].zfill(3),
                    "image_path": r["image_file"],
                    "user_text": r["user_text"],
                }
            )


def _total_formula(row_idx: int) -> str:
    """Weighted 1–5 → 0–100. Score cols start at column K (11) in HEADERS."""
    first_score_col = HEADERS.index(SCORE_COLS[0]) + 1
    parts: list[str] = []
    for i, col_name in enumerate(SCORE_COLS):
        letter = _col_letter(first_score_col + i)
        w = WEIGHTS[col_name]
        parts.append(f"{w}*({letter}{row_idx}-1)/4")
    inner = " + ".join(parts)
    return f"=ROUND(100*({inner}),1)"


def write_workbook(rows: list[dict[str, str]], out_xlsx: Path) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.datavalidation import DataValidation
    except ImportError as e:
        raise SystemExit(
            "Install openpyxl: pip install openpyxl"
        ) from e

    wb = Workbook()
    ws = wb.active
    ws.title = "grading"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for c, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    for r_i, row in enumerate(rows, start=2):
        for c, h in enumerate(HEADERS, 1):
            if h == "total_score_0to100":
                ws.cell(row=r_i, column=c, value=_total_formula(r_i))
            else:
                ws.cell(row=r_i, column=c, value=row.get(h, ""))

    # 1–5 validation on score columns
    score_start = HEADERS.index(SCORE_COLS[0]) + 1
    score_end = HEADERS.index(SCORE_COLS[-1]) + 1
    col_start = get_column_letter(score_start)
    col_end = get_column_letter(score_end)
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1="1",
        formula2="5",
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Score 1–5",
        error="Enter a number from 1 to 5 (half points OK).",
    )
    dv.add(f"{col_start}2:{col_end}{len(rows) + 1}")
    ws.add_data_validation(dv)

    yn_col = get_column_letter(HEADERS.index("would_post_YN") + 1)
    dv_yn = DataValidation(type="list", formula1='"Y,N"', allow_blank=True)
    dv_yn.add(f"{yn_col}2:{yn_col}{len(rows) + 1}")
    ws.add_data_validation(dv_yn)

    # Rubric reference sheet
    rub = wb.create_sheet("rubric")
    rub.append(["Dimension", "Weight", "1 (bad)", "3 (ok)", "5 (good)"])
    rubric_rows = [
        ("mood_fit", 0.25, "opposite mood", "somewhat fits", "clearly matches"),
        ("scene_fit", 0.15, "no relation", "generic", "evokes scene"),
        ("coherence", 0.15, "chaotic", "wobbly", "consistent"),
        ("reel_usability", 0.15, "unusable", "usable w/ trim", "ready to post"),
        ("audio_quality", 0.10, "heavy artifacts", "tolerable", "clean"),
        ("distinctiveness", 0.10, "generic", "mildly unique", "distinctive"),
        ("prompt_faithfulness", 0.10, "ignores prompt", "partial", "matches prompt"),
    ]
    for rr in rubric_rows:
        rub.append(rr)
    rub.append([])
    rub.append(["Protocol", "", "", "", ""])
    rub.append(["Raters", "1 (R1)", "", "", ""])
    rub.append(["Listen window", "8 seconds (start of clip)", "", "", ""])
    rub.append(["Volume", "fixed system volume; no EQ while scoring", "", "", ""])

    # Column widths
    widths = {
        "picture_no": 10,
        "image_file": 14,
        "user_text": 28,
        "scene_caption": 32,
        "music_prompt": 40,
        "output_audio": 22,
        "notes": 24,
    }
    for c, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 12)

    ws.freeze_panes = "A2"
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)


def merge_results(workbook: Path, results_csv: Path) -> int:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise SystemExit("pip install openpyxl") from e

    by_id: dict[str, dict] = {}
    with results_csv.open(newline="", encoding="utf-8") as f:
        for row in csv.DictReader(f):
            sid = (row.get("sample_id") or "").strip()
            stem = sid.lstrip("0") or "0"
            by_id[stem] = row
            by_id[sid] = row

    wb = load_workbook(workbook)
    ws = wb["grading"]
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    updated = 0
    for r in range(2, ws.max_row + 1):
        pic = str(ws.cell(r, col["picture_no"]).value or "").strip()
        if not pic:
            continue
        res = by_id.get(pic) or by_id.get(pic.zfill(3))
        if not res:
            continue
        if res.get("scene_caption", "").startswith("ERROR"):
            continue
        ws.cell(r, col["scene_caption"], res.get("scene_caption", ""))
        ws.cell(r, col["target_mood"], res.get("mood", ""))
        ws.cell(r, col["mood_score"], res.get("mood_score", ""))
        ws.cell(r, col["music_prompt"], res.get("music_prompt", ""))
        audio = res.get("audio_path") or ""
        if audio:
            p = Path(audio)
            ws.cell(r, col["output_audio"], f"batch_results/audio/{p.name}")
        updated += 1

    wb.save(workbook)
    return updated


def cmd_create(args: argparse.Namespace) -> None:
    images_dir = Path(args.images_dir).resolve()
    out_dir = Path(args.out_dir).resolve()
    images = discover_images(images_dir, count=args.count)
    if not images:
        raise SystemExit(f"No numbered images in {images_dir}")

    texts = load_texts_csv(Path(args.texts)) if args.texts else None
    rows = rows_from_images(
        images,
        texts,
        rater_id=args.rater,
        listen_seconds=args.listen_seconds,
        audio_subdir=args.audio_subdir,
    )

    manifest = out_dir / args.manifest_name
    workbook = out_dir / args.workbook_name
    write_manifest(rows, images_dir, manifest)
    write_workbook(rows, workbook)

    print(f"Images:  {len(images)} in {images_dir}")
    print(f"Manifest: {manifest}")
    print(f"Workbook: {workbook}")
    print(f"Rater: {args.rater} | listen: {args.listen_seconds}s")
    print("\nNext:")
    print(
        f"  python batch_eval.py --manifest {manifest} "
        f"--data-dir {images_dir.parent} --out {out_dir / 'batch_results'} "
        f"--music --music-seconds {args.listen_seconds}"
    )
    print(
        f"  python grading_workbook.py merge --workbook {workbook} "
        f"--results {out_dir / 'batch_results' / 'results.csv'}"
    )


def cmd_merge(args: argparse.Namespace) -> None:
    n = merge_results(Path(args.workbook).resolve(), Path(args.results).resolve())
    print(f"Updated {n} rows in {args.workbook}")


def main() -> None:
    ap = argparse.ArgumentParser(description="VibeSound grading Excel + manifest")
    sub = ap.add_subparsers(dest="cmd", required=True)

    c = sub.add_parser("create", help="Build manifest + grading xlsx from images/")
    c.add_argument("--images-dir", required=True, help="Folder with 1.jpg, 2.png, …")
    c.add_argument("--out-dir", default=".", help="Where to write csv/xlsx")
    c.add_argument("--count", type=int, default=50, help="Max images (0=all)")
    c.add_argument(
        "--texts",
        default="",
        help="Optional CSV: picture_no,user_text",
    )
    c.add_argument("--rater", default="R1")
    c.add_argument("--listen-seconds", type=int, default=8)
    c.add_argument("--audio-subdir", default="audio")
    c.add_argument("--manifest-name", default="batch_manifest.csv")
    c.add_argument("--workbook-name", default="vibesound_batch_grading.xlsx")
    c.set_defaults(func=cmd_create)

    m = sub.add_parser("merge", help="Fill pipeline columns from batch_eval results.csv")
    m.add_argument("--workbook", required=True)
    m.add_argument("--results", required=True)
    m.set_defaults(func=cmd_merge)

    args = ap.parse_args()
    if args.cmd == "create" and args.count == 0:
        args.count = 0  # discover all
    elif args.cmd == "create" and args.count < 0:
        args.count = 50
    args.func(args)


if __name__ == "__main__":
    main()
